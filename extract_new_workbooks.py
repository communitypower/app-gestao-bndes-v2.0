from pathlib import Path
import json
from openpyxl import load_workbook

SOURCES = {
    "atividades-grupos": Path("/home/ubuntu/upload/pasted_file_RWdb5F_Atividades-Grupos.xlsm"),
    "cronograma-r1-r2-26-agosto": Path("/home/ubuntu/upload/pasted_file_ujk65U_Cronograma-R1-e-R2-26_agosto.xlsm"),
}
OUTPUT = Path("docs/source/revisao-anexos-2026-08-30")


def cell_value(value):
    if value is None:
        return ""
    return str(value).replace("\r", " ").replace("\n", " ").strip()


def main():
    OUTPUT.mkdir(parents=True, exist_ok=True)
    catalog = []
    for label, source in SOURCES.items():
        workbook = load_workbook(source, read_only=True, data_only=False, keep_vba=True)
        summary = [f"# {label}", "", f"Fonte: `{source.name}`", ""]
        catalog.append({"label": label, "file": source.name, "sheets": workbook.sheetnames})
        for sheet in workbook.worksheets:
            structured_rows = []
            summary.extend([
                f"## {sheet.title}",
                "",
                f"Dimensões declaradas: {sheet.max_row} linhas × {sheet.max_column} colunas.",
                "",
                "| Linha | Valores não vazios |",
                "|---:|---|",
            ])
            for row_index, row in enumerate(sheet.iter_rows(max_row=sheet.max_row, values_only=True), start=1):
                values = [cell_value(value) for value in row]
                structured_rows.append({
                    "row": row_index,
                    "cells": {
                        chr(65 + column_index): value
                        for column_index, value in enumerate(values)
                        if value
                    },
                })
                compact = " · ".join(value for value in values if value)
                if compact:
                    summary.append(f"| {row_index} | {compact.replace('|', '/')} |")
            summary.append("")
            (OUTPUT / f"{label}-{sheet.title.replace(' ', '-')}.json").write_text(
                json.dumps(structured_rows, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
        (OUTPUT / f"{label}.md").write_text("\n".join(summary), encoding="utf-8")

    index = ["# Inventário dos anexos XLSM", "", "| Referência | Arquivo | Planilhas |", "|---|---|---|"]
    for item in catalog:
        index.append(f"| {item['label']} | {item['file']} | {', '.join(item['sheets'])} |")
    (OUTPUT / "indice.md").write_text("\n".join(index) + "\n", encoding="utf-8")


if __name__ == "__main__":
    main()
