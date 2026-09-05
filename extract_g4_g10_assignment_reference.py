import json
import re
from pathlib import Path

from openpyxl import load_workbook


SOURCE = Path("/home/ubuntu/upload/pasted_file_RWdb5F_Atividades-Grupos.xlsm")
OUTPUT = Path("docs/source/revisao-anexos-2026-08-30/referencia-atribuicoes-g4-g10.json")


def value(cell):
    return "" if cell is None else str(cell).replace("\n", " ").strip()


def main():
    workbook = load_workbook(SOURCE, read_only=True, data_only=False, keep_vba=True)
    worksheet = workbook["Atividades"]
    current_tome = None
    references = []

    for row in worksheet.iter_rows(values_only=True):
        title = value(row[0] if len(row) > 0 else None)
        group_reference = value(row[3] if len(row) > 3 else None)

        tome_match = re.match(r"Tomo\s+(I{1,3}|IV)\b", title, re.IGNORECASE)
        if tome_match:
            current_tome = f"Tomo {tome_match.group(1).upper()}"
            continue

        section_match = re.match(r"^(\d+)\.(\d+)\.\s+(.+)$", title)
        if not section_match or not current_tome:
            continue

        group_codes = re.findall(r"G(?:4|10)\b", group_reference)
        if not group_codes:
            continue

        roman = {"Tomo I": "I", "Tomo II": "II", "Tomo III": "III", "Tomo IV": "IV"}[current_tome]
        references.append({
            "planCode": f"{roman}.{section_match.group(1)}.{section_match.group(2)}",
            "title": section_match.group(3),
            "groupCodes": sorted(set(group_codes), key=lambda code: int(code[1:])),
            "source": "Atividades-Grupos.xlsm / aba Atividades",
        })

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(references, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({"references": len(references), "output": str(OUTPUT)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
