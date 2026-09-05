from pathlib import Path
from openpyxl import load_workbook

source_path = Path("/home/ubuntu/upload/Cronograma-R1-e-R2-26_agosto.xlsm")
output_path = Path("/home/ubuntu/estudo-bndes-gestao/docs/source/cronograma-r1-r2-26-agosto-periodos.txt")

workbook = load_workbook(source_path, read_only=False, data_only=False, keep_vba=True)
lines = []

for worksheet in workbook.worksheets:
    lines.append(f"# {worksheet.title}")
    for row_index in range(1, worksheet.max_row + 1):
        label = worksheet.cell(row=row_index, column=1).value
        if not label:
            continue
        marked_columns = []
        for column_index in range(2, min(12, worksheet.max_column + 1)):
            cell = worksheet.cell(row=row_index, column=column_index)
            fill = cell.fill
            fill_color = fill.fgColor.rgb or fill.fgColor.indexed or fill.fgColor.theme
            if fill.fill_type:
                marked_columns.append(
                    f"{cell.column_letter}:style{cell.style_id}:fill{fill_color}:value{cell.value or ''}"
                )
        if marked_columns:
            lines.append(f"{row_index:03d} | {label} | {' ; '.join(marked_columns)}")

output_path.write_text("\n".join(lines), encoding="utf-8")
print(f"Resumo salvo em {output_path}")
