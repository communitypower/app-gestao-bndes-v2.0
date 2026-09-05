from pathlib import Path
import json
from openpyxl import load_workbook

source_path = Path("/home/ubuntu/upload/Cronograma-R1-e-R2-26_agosto.xlsm")
output_path = Path("/home/ubuntu/estudo-bndes-gestao/docs/source/cronograma-r1-r2-26-agosto-estilos.json")

workbook = load_workbook(source_path, read_only=False, data_only=False, keep_vba=True)
result = []

for worksheet in workbook.worksheets:
    sheet_rows = []
    for row_number, row in enumerate(worksheet.iter_rows(), start=1):
        label = row[0].value if row else None
        markers = []
        for cell in row[1:]:
            fill = cell.fill
            fill_color = fill.fgColor.rgb or fill.fgColor.indexed or fill.fgColor.theme
            style_id = getattr(cell, "style_id", getattr(cell, "_style_id", 0))
            if cell.value is not None or style_id != 0 or fill.fill_type:
                markers.append({
                    "column": cell.column_letter,
                    "value": str(cell.value) if cell.value is not None else "",
                    "style": style_id,
                    "fill": str(fill_color) if fill_color is not None else "",
                    "fill_type": fill.fill_type or "",
                })
        if label is not None or markers:
            sheet_rows.append({"row": row_number, "label": str(label or ""), "markers": markers})
    result.append({"sheet": worksheet.title, "rows": sheet_rows})

output_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
print(f"Análise salva em {output_path}")
