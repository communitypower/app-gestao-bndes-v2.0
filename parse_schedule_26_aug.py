from pathlib import Path
from datetime import date
import json
import re
from openpyxl import load_workbook

source_path = Path("/home/ubuntu/upload/Cronograma-R1-e-R2-26_agosto.xlsm")
output_path = Path("/home/ubuntu/estudo-bndes-gestao/docs/source/cronograma-r1-r2-26-agosto-estruturado.json")

workbook = load_workbook(source_path, read_only=False, data_only=False, keep_vba=True)
sheet = workbook["diagnostico"]
roman_by_tomo = {"I": "I", "II": "II", "III": "III", "IV": "IV"}
current_tomo = None
current_chapter = None
items = []

for row_index in range(1, sheet.max_row + 1):
    label = str(sheet.cell(row=row_index, column=1).value or "").strip()
    tomo_match = re.match(r"Tomo\s+(I{1,3}V?|IV)\b", label, flags=re.IGNORECASE)
    if tomo_match:
        current_tomo = roman_by_tomo[tomo_match.group(1).upper()]
        current_chapter = None
        continue
    chapter_match = re.match(r"(\d+)\.\s+(.+)", label)
    if chapter_match and current_tomo:
        current_chapter = int(chapter_match.group(1))
        continue
    item_match = re.match(r"(\d+)\.(\d+)\.\s+(.+)", label)
    if not item_match or not current_tomo:
        continue

    chapter_number = int(item_match.group(1))
    subitem_number = int(item_match.group(2))
    active_months = []
    for column_index in range(2, 9):
        cell = sheet.cell(row=row_index, column=column_index)
        fill = cell.fill
        if fill.fill_type and fill.fgColor.rgb == "FFFFD966":
            active_months.append(column_index - 1)
    if not active_months:
        continue
    items.append({
        "tomo": current_tomo,
        "chapter": chapter_number,
        "subitem": subitem_number,
        "schedule_code": f"{chapter_number}.{subitem_number}",
        "detail_code": f"{current_tomo}.{chapter_number}.{subitem_number}",
        "title": item_match.group(3).strip(),
        "start_month": min(active_months),
        "end_month": max(active_months),
    })

output_path.write_text(json.dumps(items, ensure_ascii=False, indent=2), encoding="utf-8")
print(f"Extraídos {len(items)} subitens com período da planilha diagnostico.")
