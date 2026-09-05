from __future__ import annotations

import csv
import hashlib
import json
import re
from pathlib import Path

from openpyxl import load_workbook

UPLOAD = Path("/home/ubuntu/upload")
XLSX = UPLOAD / "matriz_atividades_relatorio1_portal_naval.xlsx"
CSV = UPLOAD / "atividades_relatorio1_portal_naval.csv"
PLAN = UPLOAD / "PlanodeatividadesparadesenvolvimentoedivulgaçãodoRelatório1.md"
OUT_JSON = Path("/home/ubuntu/portal_attachments_audit.json")
OUT_MD = Path("/home/ubuntu/portal_attachments_audit.md")


def normalized(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def digest(record: dict[str, str]) -> str:
    payload = "\n".join(f"{key}={normalized(value)}" for key, value in sorted(record.items()))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def read_csv_rows() -> tuple[list[str], list[dict[str, str]]]:
    raw_text = CSV.read_text(encoding="utf-8-sig")
    ids = re.findall(r'(?m)^"?([ABCD]\d{2});', raw_text)
    # O arquivo recebido encapsula cada linha inteira entre aspas e inclui
    # campos multilinha. A planilha contém os mesmos 36 IDs com estrutura
    # tabular válida; aqui o CSV é usado como verificação de cobertura.
    return ["id"], [{"id": record_id} for record_id in ids]


def read_workbook() -> dict[str, dict[str, object]]:
    workbook = load_workbook(XLSX, data_only=True, read_only=True)
    results: dict[str, dict[str, object]] = {}
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        header_index = next(
            (
                index
                for index, row in enumerate(rows)
                if any(normalized(value) for value in row)
            ),
            None,
        )
        if header_index is None:
            results[sheet.title] = {"header": [], "rows": []}
            continue
        header = [normalized(value) for value in rows[header_index]]
        while header and not header[-1]:
            header.pop()
        output_rows = []
        for row in rows[header_index + 1 :]:
            record = {
                header[index]: normalized(row[index]) if index < len(row) else ""
                for index in range(len(header))
            }
            if any(record.values()):
                output_rows.append(record)
        results[sheet.title] = {"header": header, "rows": output_rows}
    return results


def read_plan_ids() -> list[dict[str, str]]:
    text = PLAN.read_text(encoding="utf-8")
    records: list[dict[str, str]] = []
    for line in text.splitlines():
        if not line.startswith("|") or "**" not in line:
            continue
        cells = [normalized(cell) for cell in line.strip().strip("|").split("|")]
        if len(cells) < 8:
            continue
        match = re.fullmatch(r"\*\*(A\d{2}|B\d{2}|C\d{2}|D\d{2})\*\*", cells[0])
        if not match:
            continue
        records.append(
            {
                "id": match.group(1),
                "tomo_secao": cells[1],
                "atividade": cells[2],
                "responsavel_principal": cells[4],
                "apoio": cells[5],
            }
        )
    return records


def main() -> None:
    csv_header, csv_rows = read_csv_rows()
    workbook = read_workbook()
    activity_rows = workbook["Atividades"]["rows"]
    csv_by_id = {row["id"]: row for row in csv_rows}
    xlsx_by_id = {row["ID"]: row for row in activity_rows}
    comparable = {
        "id": "ID",
        "tomo_secao": "Tomo / seção",
        "atividade": "Atividade",
        "descricao": "Descrição pública e técnica",
        "responsavel_principal": "Responsável principal",
        "apoio": "Apoio",
        "entrega_portal": "Entrega para o Portal Naval",
        "dependencias": "Dependências",
        "palavras_chave": "Palavras-chave",
        "status": "Status",
        "tipo_conteudo": "Tipo de conteúdo",
        "visibilidade": "Visibilidade",
        "criterio_aceite": "Critério de aceite",
        "fonte_base": "Fonte-base",
    }
    differences: list[dict[str, str]] = []
    for record_id in sorted(set(csv_by_id) | set(xlsx_by_id)):
        left, right = csv_by_id.get(record_id), xlsx_by_id.get(record_id)
        if not left or not right:
            differences.append({"id": record_id, "field": "presence", "csv": bool(left), "xlsx": bool(right)})
            continue
        for csv_key, xlsx_key in comparable.items():
            if csv_key not in left:
                continue
            if normalized(left.get(csv_key)) != normalized(right.get(xlsx_key)):
                differences.append(
                    {
                        "id": record_id,
                        "field": csv_key,
                        "csv": left.get(csv_key, ""),
                        "xlsx": right.get(xlsx_key, ""),
                    }
                )
    plan_rows = read_plan_ids()
    plan_by_id = {row["id"]: row for row in plan_rows}
    summary = {
        "source_files": {
            "xlsx": str(XLSX),
            "csv": str(CSV),
            "plan": str(PLAN),
        },
        "csv": {
            "columns": csv_header,
            "row_count": len(csv_rows),
            "ids": [row["id"] for row in csv_rows],
            "unique_hashes": len({digest(row) for row in csv_rows}),
        },
        "xlsx": {
            sheet: {"row_count": len(data["rows"]), "columns": data["header"]}
            for sheet, data in workbook.items()
        },
        "markdown_plan": {
            "row_count": len(plan_rows),
            "ids": [row["id"] for row in plan_rows],
        },
        "cross_file": {
            "csv_xlsx_difference_count": len(differences),
            "csv_xlsx_differences": differences,
            "ids_only_in_csv": sorted(set(csv_by_id) - set(xlsx_by_id)),
            "ids_only_in_xlsx": sorted(set(xlsx_by_id) - set(csv_by_id)),
            "ids_only_in_plan": sorted(set(plan_by_id) - set(csv_by_id)),
            "ids_missing_in_plan": sorted(set(csv_by_id) - set(plan_by_id)),
        },
        "activities": activity_rows,
        "responsaveis": workbook["Responsáveis"]["rows"],
        "campos_portal": workbook["Campos do Portal"]["rows"],
    }
    OUT_JSON.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    lines = [
        "# Inventário dos anexos para atualização do portal",
        "",
        f"- CSV: **{len(csv_rows)}** atividades e **{len(csv_header)}** campos úteis.",
        f"- Planilha: **{len(activity_rows)}** atividades na aba `Atividades`.",
        f"- Plano em Markdown: **{len(plan_rows)}** atividades identificadas.",
        f"- Diferenças CSV × planilha: **{len(differences)}**.",
        "",
        "| ID | Tomo / seção | Responsável funcional | Apoios |",
        "|---|---|---|---|",
    ]
    for row in activity_rows:
        lines.append(
            f"| {row['ID']} | {row['Tomo / seção']} | {row['Responsável principal']} | {row['Apoio']} |"
        )
    OUT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


if __name__ == "__main__":
    main()
