from pathlib import Path
import json
from openpyxl import load_workbook

source_path = Path("/home/ubuntu/upload/pasted_file_T4R7La_Cronograma-R1-e-R2-26_agosto.xlsm")
output_dir = Path("/home/ubuntu/estudo-bndes-gestao/docs/source")
output_dir.mkdir(parents=True, exist_ok=True)

workbook = load_workbook(source_path, read_only=True, data_only=False, keep_vba=True)
result = []

for worksheet in workbook.worksheets:
    rows = []
    for row in worksheet.iter_rows(values_only=False):
        values = []
        for cell in row:
            value = cell.value
            if hasattr(value, "isoformat"):
                value = value.isoformat()
            values.append("" if value is None else str(value))
        rows.append(values)
    while rows and not any(rows[-1]):
        rows.pop()
    result.append({"name": worksheet.title, "rows": rows})

(output_dir / "cronograma-r1-r2-26-agosto.json").write_text(
    json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8"
)

for sheet in result:
    safe_name = "".join(character.lower() if character.isalnum() else "-" for character in sheet["name"]).strip("-")
    lines = [f"{index:03d} | {' | '.join(row)}" for index, row in enumerate(sheet["rows"], start=1)]
    (output_dir / f"cronograma-{safe_name}.txt").write_text("\n".join(lines), encoding="utf-8")

print("Extraídas " + str(len(result)) + " planilha(s): " + ", ".join(sheet["name"] for sheet in result))
